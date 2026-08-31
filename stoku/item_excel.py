"""
Excel template / bulk import helpers for stock items (bidhaa + bidhaa_stoku).
Column layout mirrors the manual registration form (ongezaBidhaa).
"""
import datetime
import os
import re
import time
import zipfile
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.conf import settings
from django.contrib.auth.models import User
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db.models import F, Max, Q
from django.utils import timezone

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None
    Font = None
    get_column_letter = None
    DataValidation = None

from management.models import (
    UserExtend,
    ainaMama,
    bidhaa,
    bidhaa_aina,
    bidhaa_stoku,
    ColorChange,
    mahitaji,
    makampuni,
    picha_bidhaa,
    picha_yenyewe,
    productChangeRecord,
    color_produ,
    produ_colored,
    produ_size,
    sizes,
    SizeChange,
    stokAdjustment,
    wasambazaji,
)

from .variant_utils import variant_is_color_mode

# Stable keys used in template header row and import JSON
ITEM_EXCEL_FIELDS = [
    'jina_la_bidhaa',
    'namba',
    'maelezo',
    'kundi',
    'aina',
    'chapa',
    'wasambazaji',
    'vipimo_jum',
    'vipimo_reja',
    'uwiano',
    'idadi_jumla',
    'idadi_reja',
    'bei_kununua',
    'bei_kuuza',
    'bei_kuuza_jum',
    'barcode',
    'huduma',
    'kielelezo',
    'rangi_model',
    'size',
    'picha',
]

ITEM_EXCEL_HEADERS_SWA = {
    'jina_la_bidhaa': 'Jina la Bidhaa*',
    'namba': 'Namba',
    'maelezo': 'Maelezo',
    'kundi': 'Kundi',
    'aina': 'Aina',
    'chapa': 'Chapa',
    'wasambazaji': 'Wasambazaji',
    'vipimo_jum': 'Vipimo Jumla*',
    'vipimo_reja': 'Vipimo Reja',
    'uwiano': 'Uwiano (vipimo kwa jumla)',
    'idadi_jumla': 'Idadi Jumla',
    'idadi_reja': 'Idadi Reja',
    'bei_kununua': 'Bei Kununua*',
    'bei_kuuza': 'Bei Kuuza',
    'bei_kuuza_jum': 'Bei Kuuza Jumla',
    'barcode': 'Bar Code',
    'huduma': 'Huduma (N/Y)',
    'kielelezo': 'Kielelezo (Rangi/Nyingine)',
    'rangi_model': 'Rangi/Model',
    'size': 'Size/Ukubwa',
    'picha': 'Picha (jina la faili)',
}

ITEM_EXCEL_HEADERS_ENG = {
    'jina_la_bidhaa': 'Item Name*',
    'namba': 'Item Number',
    'maelezo': 'Description',
    'kundi': 'Parent Group',
    'aina': 'Category',
    'chapa': 'Brand',
    'wasambazaji': 'Supplier',
    'vipimo_jum': 'Wholesale Unit*',
    'vipimo_reja': 'Retail Unit',
    'uwiano': 'Units per wholesale',
    'idadi_jumla': 'Wholesale Qty',
    'idadi_reja': 'Retail Qty',
    'bei_kununua': 'Purchase Price*',
    'bei_kuuza': 'Retail Sell Price',
    'bei_kuuza_jum': 'Wholesale Sell Price',
    'barcode': 'Barcode',
    'huduma': 'Service (Y/N)',
    'kielelezo': 'Variant (Color/Other)',
    'rangi_model': 'Color/Model',
    'size': 'Size',
    'picha': 'Image (filename)',
}

HEADER_ALIASES = {}
for key in ITEM_EXCEL_FIELDS:
    aliases = {key, key.replace('_', ' ')}
    swa = ITEM_EXCEL_HEADERS_SWA.get(key, '')
    eng = ITEM_EXCEL_HEADERS_ENG.get(key, '')
    if swa:
        aliases.add(swa.lower().strip())
        aliases.add(swa.lower().replace('*', '').strip())
    if eng:
        aliases.add(eng.lower().strip())
        aliases.add(eng.lower().replace('*', '').strip())
    HEADER_ALIASES[key] = aliases

KIELELEZO_DROPDOWN_SWA = ('Rangi', 'Nyingine')
KIELELEZO_DROPDOWN_ENG = ('Color', 'Other')


EXAMPLE_ROW = {
    'jina_la_bidhaa': 'Mfano: Clutch Plate',
    'namba': 'ITM-001',
    'maelezo': 'Maelezo ya bidhaa',
    'kundi': 'Vifaa vya gari',
    'aina': 'Spare parts',
    'chapa': 'Toyota',
    'wasambazaji': '',
    'vipimo_jum': 'box',
    'vipimo_reja': 'pc',
    'uwiano': 1,
    'idadi_jumla': 0,
    'idadi_reja': 10,
    'bei_kununua': 25000,
    'bei_kuuza': 35000,
    'bei_kuuza_jum': 35000,
    'barcode': '',
    'huduma': 'N',
    'kielelezo': 'Nyingine',
    'rangi_model': 'Red',
    'size': 'M',
    'picha': 'clutch_red.jpg',
}

EXAMPLE_VARIANT_ROW = {
    'jina_la_bidhaa': '',
    'namba': '',
    'maelezo': '',
    'kundi': '',
    'aina': '',
    'chapa': '',
    'wasambazaji': '',
    'vipimo_jum': '',
    'vipimo_reja': '',
    'uwiano': '',
    'idadi_jumla': '',
    'idadi_reja': 5,
    'bei_kununua': '',
    'bei_kuuza': '',
    'bei_kuuza_jum': '',
    'barcode': '',
    'huduma': '',
    'kielelezo': '',
    'rangi_model': 'Blue',
    'size': 'L',
    'picha': 'clutch_blue.jpg',
}

INSTRUCTIONS_SWA = [
    'Safu zenye * ni lazima: Jina la Bidhaa, Vipimo Jumla, Bei Kununua',
    'Kundi na Aina - majina mapya yataundwa kiotomatiki ikiwa hayapo kwenye mfumo',
    'Chapa - jina jipya litaundwa kiotomatiki; Wasambazaji lazima yawe yamesajiliwa ikiwa utayaandika',
    'Uwiano = idadi ya vipimo reja ndani ya kila jumla (mf. 1 box = 12 pc, uwiano 12)',
    'Idadi Jumla + Idadi Reja = jumla ya stock (kama fomu ya usajiri)',
    'Huduma: N = bidhaa, Y = huduma',
    'Usifute mistari miwili ya kwanza (majina ya safu na field keys)',
    'Bidhaa zenye jina lililopo tayari hazitaongezwa tena',
    'Rangi/Model na Size: mstari wa kwanza wa bidhaa umejaza taarifa zote; mistari inayofuata jaza Rangi/Model na/au Size tu',
    'Picha: andika jina la faili (mf. bidhaa.jpg). Pakia .zip yenye Excel + folda picha/ ikiwa unatumia picha',
    'Kielelezo: chagua Rangi (variants za rangi) au Nyingine (Model, Version, Size, n.k.) kutoka dropdown',
]

INSTRUCTIONS_ENG = [
    'Required columns (*): Item Name, Wholesale Unit, Purchase Price',
    'Group & Category - new names are created automatically if not in the system',
    'Brand - new names are created automatically; Supplier must exist if provided',
    'Units per wholesale = retail units inside one wholesale unit',
    'Wholesale Qty + Retail Qty = total stock (same as registration form)',
    'Service: N = goods, Y = service',
    'Do not delete the first two rows (column labels and field keys)',
    'Duplicate item names will be skipped',
    'Color/Model and Size: first row has full item details; following rows only fill Color/Model and/or Size',
    'Image: enter filename (e.g. item.jpg). Upload a .zip with Excel + picha/ folder when using images',
    'Variant label: choose Color or Other (Model, Version, Size, etc.) from the dropdown',
]


def build_ai_items_format_prompt(lang_swa=True, business_description=''):
    """
    Build a copy-paste prompt for AI tools to convert a messy product list
    into the TheBiashara items Excel template layout.
    """
    headers = _headers(lang_swa)
    header_labels = [headers[f] for f in ITEM_EXCEL_FIELDS]
    field_keys = list(ITEM_EXCEL_FIELDS)
    business_description = (business_description or '').strip()

    if lang_swa:
        parts = [
            'Wewe ni msaidizi wa kuandaa orodha ya bidhaa kwa ajili ya kuingiza kwenye mfumo wa TheBiashara.',
            '',
            'LENGO:',
            'Nipe orodha yangu ya bidhaa (ambayo inaweza kuwa Excel, Word, PDF, au maandishi yasiyo na muundo).',
            'Ibadilishe iwe jedwali linalolingana NA template ya Excel ya bidhaa ya TheBiashara ili niweze kuipakia moja kwa moja.',
            '',
        ]
        if business_description:
            parts.extend([
                'MAELEZO YA BIASHARA YANGU (tumia kuelewa aina ya bidhaa na kupanga vikundi):',
                business_description,
                '',
            ])
        else:
            parts.extend([
                'MAELEZO YA BIASHARA:',
                '(Hakuna maelezo yaliyotolewa. Kadiria aina ya biashara kutoka majina ya bidhaa.)',
                '',
            ])

        parts.extend([
            'MUUNDO WA MATOKEO (LAZIMA):',
            '1. Mstari wa 1: vichwa vya safu kwa lugha ya kawaida, kwa mpangilio huu haswa:',
            ' | '.join(header_labels),
            '2. Mstari wa 2: field keys kwa mpangilio huu haswa (usibadilishe):',
            ', '.join(field_keys),
            '3. Kuanzia mstari wa 3: bidhaa moja (au variant) kwa kila mstari.',
            '',
            'SAFU ZINAZOTAKIWA (*):',
            '- Jina la Bidhaa*',
            '- Vipimo Jumla*',
            '- Bei Kununua*',
            '',
            'SAFU ZA HIARI (zinaweza kukosekana kwenye orodha yangu):',
            '- Kundi (Parent Group), Aina (Category), Chapa (Brand), Wasambazaji, Namba, Maelezo, Vipimo Reja, Uwiano, Idadi, Bei za kuuza, Barcode, Huduma, Kielelezo, Rangi/Model, Size, Picha.',
            '',
            'SHERIA ZA KUPANGA:',
            '1. Ikiwa orodha yangu HAINA Kundi / Aina / Chapa, ZITENGENEZE wewe kwa mantiki kulingana na majina ya bidhaa na maelezo ya biashara.',
            '2. Weka bidhaa zinazofanana chini ya Kundi na Aina moja; Chapa iwe jina la brand ikiwa linajulikana, vinginevyo "Generic" au chapa inayofaa.',
            '3. Usibuni barcode isipokuwa ipo kwenye orodha yangu.',
            '4. Huduma: andika N kwa bidhaa za kawaida, Y kwa huduma.',
            '5. Kielelezo: tumia "Rangi" au "Nyingine" tu.',
            '6. Variants: mstari wa kwanza wa bidhaa uwe na taarifa zote; mistari inayofuata ya rangi/size ijaze Rangi/Model na/au Size tu, acha safu nyingine tupu.',
            '7. Hifadhi bei na idadi kutoka orodha yangu pale zinapopatikana. Bei Kununua ikikosekana tumia 0.',
            '8. Vipimo Jumla ikikosekana tumia "pc" au kipimo kinachofaa.',
            '9. Usiongeze safu zingine; usibadilishe majina ya vichwa wala field keys.',
            '10. Toa matokeo kama jedwali (CSV/TSV au Excel) tayari kunakiliwa — bila maelezo ya ziada isipokuwa ninavyouliza.',
            '',
            'ORODHA YANGU YA BIDHAA:',
            '(Bandika hapa orodha yako / attach file yako)',
        ])
        return '\n'.join(parts)

    parts = [
        'You are helping prepare a product list for import into TheBiashara inventory system.',
        '',
        'GOAL:',
        'I will give you my product list (Excel, Word, PDF, or unstructured text).',
        'Convert it into a table that EXACTLY matches TheBiashara items Excel template so I can upload it directly.',
        '',
    ]
    if business_description:
        parts.extend([
            'MY BUSINESS CONTEXT (use this to understand products and invent sensible groupings):',
            business_description,
            '',
        ])
    else:
        parts.extend([
            'BUSINESS CONTEXT:',
            '(None provided. Infer the business type from the product names.)',
            '',
        ])

    parts.extend([
        'REQUIRED OUTPUT FORMAT:',
        '1. Row 1: human-readable column headers in this exact order:',
        ' | '.join(header_labels),
        '2. Row 2: field keys in this exact order (do not change):',
        ', '.join(field_keys),
        '3. From row 3 onward: one product (or variant) per row.',
        '',
        'REQUIRED COLUMNS (*):',
        '- Item Name*',
        '- Wholesale Unit*',
        '- Purchase Price*',
        '',
        'OPTIONAL COLUMNS (my list may not have these):',
        '- Parent Group, Category, Brand, Supplier, Item Number, Description, Retail Unit, Units per wholesale, Quantities, Sell prices, Barcode, Service, Variant, Color/Model, Size, Image.',
        '',
        'FORMATTING RULES:',
        '1. If my list has NO Parent Group / Category / Brand, CREATE sensible ones from product names and the business context.',
        '2. Group similar items under the same Parent Group and Category; Brand should be a real brand when known, otherwise "Generic" or a fitting placeholder.',
        '3. Do not invent barcodes unless they appear in my list.',
        '4. Service: use N for goods, Y for services.',
        '5. Variant label: use only "Color" or "Other".',
        '6. Variants: first item row has full details; following color/size rows fill only Color/Model and/or Size, leave other columns blank.',
        '7. Keep prices and quantities from my list when available. If Purchase Price is missing, use 0.',
        '8. If Wholesale Unit is missing, use "pc" or another suitable unit.',
        '9. Do not add extra columns; do not rename headers or field keys.',
        '10. Return the result as a table (CSV/TSV or Excel) ready to copy — no extra commentary unless I ask.',
        '',
        'MY PRODUCT LIST:',
        '(Paste your list / attach your file here)',
    ])
    return '\n'.join(parts)


def _headers(lang_swa=True):
    return ITEM_EXCEL_HEADERS_SWA if lang_swa else ITEM_EXCEL_HEADERS_ENG


def _kielelezo_dropdown_choices(lang_swa=True):
    return KIELELEZO_DROPDOWN_SWA if lang_swa else KIELELEZO_DROPDOWN_ENG


def _kielelezo_to_dropdown(color_attr, lang_swa=True):
    """Map stored colorAttr to template dropdown label."""
    if variant_is_color_mode(color_attr):
        return _kielelezo_dropdown_choices(lang_swa)[0]
    if _cell_str(color_attr):
        return _kielelezo_dropdown_choices(lang_swa)[1]
    return ''


def _normalize_kielelezo(val):
    """Map dropdown / legacy values to bidhaa.colorAttr."""
    s = _cell_str(val).lower()
    if not s:
        return ''
    if s in ('rangi', 'color', 'colour', 'colors', 'colours'):
        return 'Rangi'
    if s in ('nyingine', 'other', 'others', 'model', 'version', 'size', 'ukubwa'):
        return 'Model'
    return _cell_str(val)


def _apply_kielelezo_dropdown(ws, lang_swa=True):
    """Add Excel dropdown for kielelezo column on data rows."""
    if not DataValidation or not get_column_letter:
        return
    choices = _kielelezo_dropdown_choices(lang_swa)
    col_letter = get_column_letter(ITEM_EXCEL_FIELDS.index('kielelezo') + 1)
    formula = '"{}"'.format(','.join(choices))
    dv = DataValidation(type='list', formula1=formula, allow_blank=True)
    dv.error = 'Chagua Rangi au Nyingine' if lang_swa else 'Choose Color or Other'
    dv.errorTitle = 'Kielelezo' if lang_swa else 'Variant label'
    dv.prompt = 'Chagua aina ya kielelezo' if lang_swa else 'Choose variant type'
    dv.promptTitle = 'Kielelezo' if lang_swa else 'Variant'
    dv.add(f'{col_letter}3:{col_letter}5000')
    ws.add_data_validation(dv)


def _example_rows_for_template(lang_swa=True):
    main = dict(EXAMPLE_ROW)
    main['kielelezo'] = _kielelezo_dropdown_choices(lang_swa)[1]
    return main, dict(EXAMPLE_VARIANT_ROW)


def _sanitize_excel_text(val):
    if val is None:
        return ''
    if isinstance(val, (int, float, Decimal)) and not isinstance(val, bool):
        return val
    s = str(val)
    s = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\uFFFE\uFFFF]', '', s)
    s = re.sub(r'[\uD800-\uDFFF]', '', s)
    return s[:32767]


def _row_values(row_dict):
    return [_sanitize_excel_text(row_dict.get(f, '')) for f in ITEM_EXCEL_FIELDS]


def build_items_xlsx_bytes(data_rows=None, lang_swa=True, include_instructions=True, include_example=True):
    """Build a valid .xlsx workbook using openpyxl."""
    if Workbook is None:
        raise RuntimeError('openpyxl is not installed')

    headers = _headers(lang_swa)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bidhaa' if lang_swa else 'Items'

    label_row = [_sanitize_excel_text(headers[f]) for f in ITEM_EXCEL_FIELDS]
    key_row = list(ITEM_EXCEL_FIELDS)
    ws.append(label_row)
    ws.append(key_row)

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    rows = data_rows or []
    if rows:
        for row in rows:
            ws.append(_row_values(row))
    elif include_example:
        example_main, example_variant = _example_rows_for_template(lang_swa)
        ws.append(_row_values(example_main))
        ws.append(_row_values(example_variant))

    _apply_kielelezo_dropdown(ws, lang_swa=lang_swa)

    if get_column_letter:
        for col_idx, field in enumerate(ITEM_EXCEL_FIELDS, start=1):
            width = max(len(str(headers[field])), len(field)) + 2
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 40)

    if include_instructions:
        inst_title = 'Maelezo' if lang_swa else 'Instructions'
        ws2 = wb.create_sheet(inst_title)
        header = 'Maelezo' if lang_swa else 'Note'
        ws2.append([header])
        ws2[1][0].font = bold
        for line in (INSTRUCTIONS_SWA if lang_swa else INSTRUCTIONS_ENG):
            ws2.append([line])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def stoku_record_to_export_row(rec, lang_swa=True):
    uwiano = max(int(float(rec.get('uwiano') or 1)), 1)
    idadi = int(rec.get('idadi') or 0)
    idadi_jum = idadi // uwiano if uwiano > 1 else 0
    idadi_rej = (idadi % uwiano) if uwiano > 1 else idadi
    return {
        'jina_la_bidhaa': rec.get('bidhaaN') or '',
        'namba': rec.get('namba') or '',
        'maelezo': rec.get('maelezo') or '',
        'kundi': rec.get('group_name') or '',
        'aina': rec.get('ainaN') or '',
        'chapa': rec.get('brand') or '',
        'wasambazaji': rec.get('vendor') or '',
        'vipimo_jum': rec.get('vipimoJum') or rec.get('vipimo') or 'pc',
        'vipimo_reja': rec.get('vipimo') or 'pc',
        'uwiano': uwiano,
        'idadi_jumla': idadi_jum,
        'idadi_reja': idadi_rej,
        'bei_kununua': float(rec.get('Bei_kununua') or 0),
        'bei_kuuza': float(rec.get('Bei_kuuza') or 0),
        'bei_kuuza_jum': float(rec.get('Bei_kuuza_jum') or 0),
        'barcode': rec.get('sirio') or '',
        'huduma': 'Y' if rec.get('service') else 'N',
        'kielelezo': _kielelezo_to_dropdown(rec.get('colorAttr'), lang_swa=lang_swa),
        'rangi_model': '',
        'size': '',
        'picha': '',
    }


def _empty_master_export_row():
    row = {f: '' for f in ITEM_EXCEL_FIELDS}
    return row


def _image_basename_for_color(color_id):
    pic = picha_bidhaa.objects.filter(color_produ_id=color_id).select_related('picha').first()
    if not pic or not pic.picha or not pic.picha.picha:
        return ''
    return os.path.basename(str(pic.picha.picha.name))


def _export_variant_lines(stoku_id, intp_id, uwiano):
    lines = []
    colors = produ_colored.objects.filter(
        bidhaa_id=stoku_id,
        Interprise_id=intp_id,
        color__colored=True,
    ).select_related('color').order_by('id')
    for pc in colors:
        size_qs = produ_size.objects.filter(
            bidhaa_id=stoku_id,
            sized__color=pc.color,
        ).select_related('sized').order_by('id')
        pic_name = _image_basename_for_color(pc.color_id)
        if size_qs.exists():
            for ps in size_qs:
                qty = float(ps.idadi or 0)
                idadi_jum = int(qty // uwiano) if uwiano > 1 else 0
                idadi_rej = int(qty % uwiano) if uwiano > 1 else int(qty)
                lines.append({
                    'rangi_model': pc.color.color_name or '',
                    'size': ps.sized.size if ps.sized else '',
                    'idadi_jumla': idadi_jum,
                    'idadi_reja': idadi_rej,
                    'picha': pic_name,
                })
                pic_name = ''
        else:
            qty = float(pc.idadi or 0)
            idadi_jum = int(qty // uwiano) if uwiano > 1 else 0
            idadi_rej = int(qty % uwiano) if uwiano > 1 else int(qty)
            lines.append({
                'rangi_model': pc.color.color_name or '',
                'size': '',
                'idadi_jumla': idadi_jum,
                'idadi_reja': idadi_rej,
                'picha': pic_name,
            })
    if not lines:
        stoku = bidhaa_stoku.objects.filter(pk=stoku_id).select_related('bidhaa').first()
        if stoku:
            pic = picha_bidhaa.objects.filter(bidhaa_id=stoku.bidhaa_id).select_related('picha').first()
            if pic and pic.picha and pic.picha.picha:
                lines.append({
                    'rangi_model': '',
                    'size': '',
                    'idadi_jumla': '',
                    'idadi_reja': '',
                    'picha': os.path.basename(str(pic.picha.picha.name)),
                })
    return lines


def fetch_export_rows(intp, filters=None, lang_swa=True):
    """Return deduplicated export rows for the current enterprise."""
    filters = filters or {}
    itm_qs = bidhaa_stoku.objects.filter(
        Q(idadi__gt=0) | Q(inapacha=False) | Q(produced__notsure=True),
        Interprise__owner=intp.owner.id,
    ).annotate(
        st=F('Interprise'),
        group_name=F('bidhaa__bidhaa_aina__mahi__mahitaji'),
        kampuni=F('bidhaa__kampuni__id'),
        aina=F('bidhaa__bidhaa_aina__id'),
        namba=F('bidhaa__namba'),
        material=F('bidhaa__material'),
        brand=F('bidhaa__kampuni_id__kampuni_jina'),
        ainaN=F('bidhaa__bidhaa_aina__aina'),
        vendor=F('msambaji_id__jina'),
        bidhaaN=F('bidhaa__bidhaa_jina'),
        maelezo=F('bidhaa__maelezo'),
        vipimo=F('bidhaa__vipimo'),
        uwiano=F('bidhaa__idadi_jum'),
        vipimoJum=F('bidhaa__vipimo_jum'),
        colorAttr=F('bidhaa__colorAttr'),
    ).filter(st=intp.id)

    records = list(itm_qs.values(
        'id', 'idadi', 'Bei_kununua', 'Bei_kuuza', 'Bei_kuuza_jum', 'sirio', 'service',
        'group_name', 'kampuni', 'aina', 'namba', 'material', 'brand', 'ainaN', 'vendor',
        'bidhaaN', 'maelezo', 'bidhaa_id', 'msambaji_id', 'vipimo', 'uwiano', 'vipimoJum',
        'colorAttr',
    ))

    flt = int(filters.get('f') or 0)
    bflt = int(filters.get('bf') or 0)
    supflt = int(filters.get('sup') or 0)
    uncat = str(filters.get('uncat') or '') in ('1', 'true', 'True')

    export_rows = []
    seen = set()
    for rec in records:
        if not ((float(rec.get('Bei_kuuza') or 0) > 0) or not rec.get('material')):
            continue
        if rec.get('service'):
            continue
        if uncat and rec.get('aina'):
            continue
        if flt > 0 and int(rec.get('aina') or 0) != flt:
            continue
        if bflt > 0 and int(rec.get('kampuni') or 0) != bflt:
            continue
        if supflt > 0 and int(rec.get('msambaji_id') or 0) != supflt:
            continue
        key = rec.get('bidhaa_id')
        if key in seen:
            continue
        seen.add(key)
        base = stoku_record_to_export_row(rec, lang_swa=lang_swa)
        if not _cell_str(base.get('jina_la_bidhaa')):
            continue
        uwiano = max(int(float(rec.get('uwiano') or 1)), 1)
        variant_lines = _export_variant_lines(rec.get('id'), intp.id, uwiano)
        if not variant_lines:
            export_rows.append(base)
            continue
        for idx, vl in enumerate(variant_lines):
            if idx == 0:
                row = dict(base)
                row.update(vl)
            else:
                row = _empty_master_export_row()
                row.update(vl)
            export_rows.append(row)
    return export_rows


def extract_import_bundle(upload):
    """Return (xlsx_file_like, image_map) from .xlsx or .zip upload."""
    name = (getattr(upload, 'name', '') or '').lower()
    image_map = {}
    if name.endswith('.zip'):
        xlsx_bytes = None
        with zipfile.ZipFile(upload) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                base = os.path.basename(info.filename)
                lower = info.filename.lower().replace('\\', '/')
                if lower.endswith('.xlsx') and not base.startswith('~'):
                    xlsx_bytes = zf.read(info.filename)
                elif lower.endswith(('.jpg', '.jpeg', '.png', '.webp', '.gif')):
                    image_map[base.lower()] = zf.read(info.filename)
        if not xlsx_bytes:
            raise ValueError('ZIP file must contain an .xlsx workbook')
        return BytesIO(xlsx_bytes), image_map
    if hasattr(upload, 'seek'):
        upload.seek(0)
    return upload, image_map


def _is_field_key_row(row):
    """True when row 2 is the internal field-key header (jina_la_bidhaa, ...)."""
    if not row:
        return False
    cells = [_cell_str(c).lower() for c in row[:len(ITEM_EXCEL_FIELDS)]]
    if cells and cells[0] == 'jina_la_bidhaa':
        return True
    matches = sum(
        1 for i, f in enumerate(ITEM_EXCEL_FIELDS)
        if i < len(cells) and cells[i] == f
    )
    return matches >= 8


def _map_row_by_headers(header_row, data_row):
    """Map a data row using human-readable header labels from row 1."""
    alias_to_field = {}
    for field in ITEM_EXCEL_FIELDS:
        alias_to_field[field.lower()] = field
        alias_to_field[field.replace('_', ' ').lower()] = field
        for headers in (ITEM_EXCEL_HEADERS_SWA, ITEM_EXCEL_HEADERS_ENG):
            label = headers.get(field, '')
            if label:
                alias_to_field[label.lower().replace('*', '').strip()] = field

    mapped = {}
    for col_idx, val in enumerate(data_row):
        if col_idx >= len(header_row):
            break
        header = _cell_str(header_row[col_idx]).lower().replace('*', '').strip()
        field = alias_to_field.get(header)
        if field and val is not None and str(val).strip() != '':
            mapped[field] = val
    return mapped


def parse_uploaded_xlsx(file_obj):
    """Parse uploaded .xlsx into list of row dicts with canonical keys."""
    if load_workbook is None:
        raise RuntimeError('openpyxl is not installed')
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return []

    header_row = all_rows[0]
    data_start_idx = 2
    if len(all_rows) >= 2 and _is_field_key_row(all_rows[1]):
        data_start_idx = 3

    rows = []
    for idx in range(data_start_idx, len(all_rows) + 1):
        row = all_rows[idx - 1]
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue

        raw = {}
        for col_idx, field in enumerate(ITEM_EXCEL_FIELDS):
            if col_idx < len(row):
                raw[field] = row[col_idx]

        header_mapped = _map_row_by_headers(header_row, row)
        merged = dict(raw)
        merged.update(header_mapped)

        normalized = normalize_import_row(merged)
        merged.update({k: v for k, v in normalized.items() if v is not None and str(v).strip() != ''})

        name = _cell_str(merged.get('jina_la_bidhaa'))
        if _is_skippable_name(name):
            continue
        has_master = bool(name)
        has_variant = bool(
            _cell_str(merged.get('rangi_model'))
            or _cell_str(merged.get('size'))
            or _cell_str(merged.get('picha'))
        )
        if not has_master and not has_variant:
            continue
        merged['_row_num'] = idx
        rows.append(merged)
    return rows


def _cell_str(val):
    if val is None:
        return ''
    return str(val).strip()


def _parse_bool_yn(val, default=False):
    s = _cell_str(val).lower()
    if not s:
        return default
    return s in ('y', 'yes', 'ndio', '1', 'true', 'huduma')


def _parse_decimal(val, default=0):
    s = _cell_str(val).replace(',', '')
    if not s:
        return Decimal(str(default))
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal(str(default))


def _parse_int(val, default=0):
    s = _cell_str(val).replace(',', '')
    if not s:
        return int(default)
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return int(default)


def normalize_import_row(raw):
    """Map a spreadsheet row dict to canonical field keys."""
    if not isinstance(raw, dict):
        return {}
    normalized = {}
    lower_map = {}
    for k, v in raw.items():
        if k is None:
            continue
        lower_map[str(k).strip().lower()] = v

    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in lower_map:
                normalized[field] = lower_map[alias]
                break
    return normalized


def _format_adj_code(adjno):
    adjno = int(adjno or 1)
    if adjno < 10:
        return '000' + str(adjno)
    if adjno < 100:
        return '00' + str(adjno)
    if adjno < 1000:
        return '0' + str(adjno)
    return str(adjno)


def _load_next_adj_num(duka):
    mx = stokAdjustment.objects.filter(Interprise=duka.id).aggregate(m=Max('code_num'))['m']
    return int(mx) if mx else 1


class _AdjSeq:
    """In-memory adjustment numbers so bulk import does not query .last() per item."""

    def __init__(self, start):
        self.n = int(start or 1)

    def next(self):
        adjno = self.n
        self.n = adjno + 1
        return _format_adj_code(adjno), self.n


def _next_adj_code(duka, adj_seq=None):
    if adj_seq is None:
        adj_seq = _AdjSeq(_load_next_adj_num(duka))
    return adj_seq.next()


def _import_ctx(duka, intp, user, ctx=None):
    if ctx is not None:
        return ctx
    return {
        'adj_seq': _AdjSeq(_load_next_adj_num(duka)),
        'owner_user': User.objects.get(pk=intp.admin),
        'op_user': UserExtend.objects.get(user=user),
        'aina': {},
        'brand': {},
        'supplier': {},
    }


def _resolve_aina(duka, aina_name, kundi_name):
    """Find or create category (+ parent group if needed)."""
    aina_name = _cell_str(aina_name)
    kundi_name = _cell_str(kundi_name)
    if not aina_name:
        return None, None

    qs = bidhaa_aina.objects.filter(Interprise=duka, aina__iexact=aina_name)
    if kundi_name:
        qs = qs.filter(mahi__mahitaji__iexact=kundi_name)
    obj = qs.first()
    if not obj:
        qs = bidhaa_aina.objects.filter(
            Interprise__owner=duka.owner.id,
            aina__iexact=aina_name,
        )
        if kundi_name:
            qs = qs.filter(mahi__mahitaji__iexact=kundi_name)
        obj = qs.first()
    if obj:
        return obj, obj.mahi

    mahi = _resolve_group(duka, kundi_name, create=True)
    if not mahi and kundi_name:
        mahi = _resolve_group(duka, kundi_name, create=True)
    if not mahi:
        return None, None

    obj = bidhaa_aina(Interprise=duka, aina=aina_name, mahi=mahi)
    obj.save()
    return obj, mahi


def _resolve_group(duka, kundi_name, create=False):
    kundi_name = _cell_str(kundi_name)
    if not kundi_name:
        return None
    obj = mahitaji.objects.filter(Interprise=duka, mahitaji__iexact=kundi_name).first()
    if not obj:
        obj = mahitaji.objects.filter(
            Interprise__owner=duka.owner.id,
            mahitaji__iexact=kundi_name,
        ).first()
    if obj or not create:
        return obj
    ainamam = ainaMama.objects.first()
    if not ainamam:
        return None
    obj = mahitaji(Interprise=duka, mahitaji=kundi_name, aina=ainamam)
    obj.save()
    return obj


def _resolve_brand(duka, name):
    name = _cell_str(name)
    if not name:
        return None
    obj = makampuni.objects.filter(Interprise=duka, kampuni_jina__iexact=name).first()
    if not obj:
        obj = makampuni.objects.filter(
            Interprise__owner=duka.owner.id,
            kampuni_jina__iexact=name,
        ).first()
    if obj:
        return obj
    obj = makampuni(Interprise=duka, kampuni_jina=name)
    obj.save()
    return obj


def _resolve_supplier(owner_user, name):
    name = _cell_str(name)
    if not name:
        return None
    return wasambazaji.objects.filter(
        owner=owner_user,
        jina__iexact=name,
    ).first()


def _is_skippable_name(name):
    s = _cell_str(name)
    if not s:
        return True
    low = s.lower()
    if low.startswith('mfano:') or low.startswith('example:'):
        return True
    if low in ('jina_la_bidhaa', 'jina la bidhaa', 'jina la bidhaa*', 'item name', 'item name*'):
        return True
    return False


def _has_variant_data(row):
    return bool(
        _cell_str(row.get('rangi_model'))
        or _cell_str(row.get('size'))
        or _cell_str(row.get('picha'))
    )


def _variant_slice(row):
    return {
        'rangi_model': _cell_str(row.get('rangi_model')),
        'size': _cell_str(row.get('size')),
        'picha': _cell_str(row.get('picha')),
        'idadi_jumla': _parse_int(row.get('idadi_jumla'), 0),
        'idadi_reja': _parse_int(row.get('idadi_reja'), 0),
    }


def group_import_rows(rows):
    """Group flat spreadsheet rows into items with variant continuation rows."""
    groups = []
    current = None
    for row in rows:
        if not isinstance(row, dict):
            continue
        row_num = row.get('_row_num', 0)
        name = _cell_str(row.get('jina_la_bidhaa'))
        if name and not _is_skippable_name(name):
            if current:
                groups.append(current)
            current = {
                'master': dict(row),
                'variants': [],
                'start_row': row_num,
            }
            if _has_variant_data(row):
                current['variants'].append(_variant_slice(row))
        elif current and _has_variant_data(row):
            current['variants'].append(_variant_slice(row))
    if current:
        groups.append(current)
    return groups


def _variant_color_code(name):
    name = _cell_str(name)
    if name.startswith('#') and len(name) in (4, 7):
        return name
    code = abs(hash(name)) % 0xFFFFFF
    return f'#{code:06x}'


def _resolve_image_bytes(filename, image_map):
    if not filename or not image_map:
        return None, None
    key = os.path.basename(filename).lower()
    if key in image_map:
        return image_map[key], os.path.basename(filename)
    return None, None


def _attach_product_image(duka, entp, produ, color_produ_obj, image_bytes, filename):
    if not image_bytes:
        return
    gcs_storage = default_storage
    if not settings.DEBUG:
        gcs_storage = settings.GCS_STORAGE_INSTANCE
    ext = filename.rsplit('.', 1)[-1] if '.' in filename else 'jpg'
    path = gcs_storage.save(
        f"products/{duka.id}_{int(time.time())}.{ext}",
        ContentFile(image_bytes),
    )
    photo = picha_yenyewe.objects.create(
        picha=path,
        pic_size=len(image_bytes),
        owner=entp.owner.user,
    )
    picha_bidhaa.objects.create(
        color_produ=color_produ_obj,
        picha=photo,
        bidhaa=produ,
    )


def _build_color_groups(variant_rows):
    by_name = {}
    order = []
    last_rangi = ''
    for v in variant_rows:
        rangi = _cell_str(v.get('rangi_model')) or last_rangi
        if _cell_str(v.get('rangi_model')):
            last_rangi = rangi
        if not rangi and not _cell_str(v.get('size')):
            if _cell_str(v.get('picha')) and order:
                by_name[order[-1]]['picha'] = v['picha']
            continue
        if not rangi:
            rangi = 'default'
        if rangi not in by_name:
            by_name[rangi] = {
                'rangi_model': rangi,
                'idadi_jumla': v.get('idadi_jumla', 0),
                'idadi_reja': v.get('idadi_reja', 0),
                'picha': v.get('picha', ''),
                'sizes': [],
            }
            order.append(rangi)
        else:
            if _cell_str(v.get('picha')):
                by_name[rangi]['picha'] = v['picha']
            if not _cell_str(v.get('size')):
                by_name[rangi]['idadi_jumla'] = v.get('idadi_jumla', 0)
                by_name[rangi]['idadi_reja'] = v.get('idadi_reja', 0)
        size = _cell_str(v.get('size'))
        if size:
            by_name[rangi]['sizes'].append({
                'size': size,
                'idadi_jumla': v.get('idadi_jumla', 0),
                'idadi_reja': v.get('idadi_reja', 0),
                'picha': v.get('picha', ''),
            })
    return [by_name[k] for k in order]


def _enrich_variants_from_master(master, variants):
    """Copy master stock qty to first variant when variant rows omit qty."""
    if not variants:
        return variants
    master_jum = _parse_int(master.get('idadi_jumla'), 0)
    master_rej = _parse_int(master.get('idadi_reja'), 0)
    if not master_jum and not master_rej:
        return variants
    first = dict(variants[0])
    if not first.get('idadi_jumla') and not first.get('idadi_reja'):
        first['idadi_jumla'] = master_jum
        first['idadi_reja'] = master_rej
        return [first] + list(variants[1:])
    return variants


def _apply_variants(duka, intp, user, produ, produ_stock, reg, uwiano, variant_rows, image_map, color_attr=''):
    color_groups = _build_color_groups(variant_rows)
    if not color_groups:
        return
    is_color_mode = variant_is_color_mode(color_attr)
    total_qty = 0
    for cg in color_groups:
        rang = color_produ()
        rang.color_code = _variant_color_code(cg['rangi_model']) if is_color_mode else '#cccccc'
        rang.color_name = cg['rangi_model']
        rang.nick_name = cg['rangi_model']
        rang.colored = True
        rang.bidhaa = produ
        rang.save()

        color = produ_colored()
        color.bidhaa = produ_stock
        color.color = rang
        color.Interprise = duka
        color.owner = duka.owner.user

        reg_c = ColorChange()
        reg_c.change = reg

        if cg['sizes']:
            idd = 0
            size_entries = []
            for s in cg['sizes']:
                szs = sizes()
                szs.size = s['size']
                szs.color = rang
                szs.save()

                szd = produ_size()
                szd.bidhaa = produ_stock
                szd.sized = szs
                szd.Interprise = duka
                szd.idadi = (float(uwiano) * float(s['idadi_jumla'])) + float(s['idadi_reja'])
                szd.owner = duka.owner.user
                szd.save()
                size_entries.append(szd)
                idd += float(szd.idadi)

                img_bytes, img_name = _resolve_image_bytes(s.get('picha'), image_map)
                if img_bytes:
                    _attach_product_image(duka, intp.Interprise, produ, rang, img_bytes, img_name)

            color.idadi = idd
            color.save()
            reg_c.color = color
            reg_c.qty = idd
            reg_c.save()
            for szd in size_entries:
                reg_s = SizeChange()
                reg_s.size = szd
                reg_s.color = reg_c
                reg_s.qty = float(szd.idadi)
                reg_s.save()
            total_qty += idd
        else:
            qty = (float(uwiano) * float(cg['idadi_jumla'])) + float(cg['idadi_reja'])
            color.idadi = qty
            color.save()
            reg_c.color = color
            reg_c.qty = qty
            reg_c.save()
            total_qty += qty

        img_bytes, img_name = _resolve_image_bytes(cg.get('picha'), image_map)
        if img_bytes:
            _attach_product_image(duka, intp.Interprise, produ, rang, img_bytes, img_name)

    if total_qty > 0:
        produ_stock.idadi = total_qty
        produ_stock.save()
        reg.qty = total_qty
        reg.save()


def _attach_item_image_no_variant(duka, intp, produ, produ_stock, picha_name, image_map):
    img_bytes, img_name = _resolve_image_bytes(picha_name, image_map)
    if not img_bytes:
        return
    hamna_rangi = color_produ()
    hamna_rangi.color_code = '#ffffff'
    hamna_rangi.color_name = 'none'
    hamna_rangi.colored = False
    hamna_rangi.bidhaa = produ
    hamna_rangi.save()

    coloring = produ_colored()
    coloring.bidhaa = produ_stock
    coloring.color = hamna_rangi
    coloring.idadi = 0
    coloring.Interprise = intp.Interprise
    coloring.owner = intp.Interprise.owner.user
    coloring.save()

    _attach_product_image(duka, intp.Interprise, produ, hamna_rangi, img_bytes, img_name)


def _existing_item_has_variants(produ_stock):
    if not produ_stock:
        return False
    return produ_colored.objects.filter(
        bidhaa=produ_stock,
        color__colored=True,
    ).exists()


def _enrich_existing_item(duka, intp, user, existing, produ_stock, row, variants, image_map):
    """Fill missing category, brand, colorAttr, and variants on an existing item."""
    updated_fields = []
    color_attr = _normalize_kielelezo(row.get('kielelezo'))
    if variants and not color_attr:
        color_attr = 'Model'

    selected_aina, mahi = _resolve_aina(duka, row.get('aina'), row.get('kundi'))
    if selected_aina and not existing.bidhaa_aina_id:
        existing.bidhaa_aina = selected_aina
        existing.Mahi = mahi
        updated_fields.append('category')

    brand = _resolve_brand(duka, row.get('chapa'))
    if brand and not existing.kampuni_id:
        existing.kampuni = brand
        updated_fields.append('brand')

    if color_attr and not _cell_str(existing.colorAttr):
        existing.colorAttr = color_attr
        updated_fields.append('colorAttr')

    if updated_fields:
        existing.save()

    variant_applied = False
    if variants and produ_stock and not _existing_item_has_variants(produ_stock):
        uwiano = max(int(float(existing.idadi_jum or 1)), 1)
        variants = _enrich_variants_from_master(row, variants)
        reg = productChangeRecord.objects.filter(prod=produ_stock).order_by('-id').first()
        if not reg:
            reg = productChangeRecord()
            reg.prod = produ_stock
            reg.qty = float(produ_stock.idadi or 0)
            reg.save()
        _apply_variants(
            duka, intp, user, existing, produ_stock, reg, uwiano, variants, image_map,
            color_attr=color_attr or existing.colorAttr or '',
        )
        variant_applied = True
        updated_fields.append('variants')

    return updated_fields, variant_applied


def import_item_group(duka, intp, user, group, image_map=None, row_num=0, ctx=None):
    """Create bidhaa + variants + images from a grouped import."""
    image_map = image_map or {}
    ctx = _import_ctx(duka, intp, user, ctx)
    row = group.get('master') or {}
    variants = group.get('variants') or []
    row_num = group.get('start_row') or row_num

    name = _cell_str(row.get('jina_la_bidhaa'))
    if _is_skippable_name(name):
        return {
            'success': False,
            'row': row_num,
            'skipped': True,
            'message_swa': 'Mstari wa kichwa au mfano — umerukwa',
            'message_eng': 'Header or example row — skipped',
        }
    if not name:
        return {
            'success': False,
            'row': row_num,
            'message_swa': 'Mstari hauna jina la bidhaa',
            'message_eng': 'Row is missing item name',
        }

    existing = bidhaa.objects.filter(bidhaa_jina=name, owner=duka.owner.user.id).first()
    if existing:
        produ_stock = bidhaa_stoku.objects.filter(
            bidhaa=existing,
            Interprise=intp.Interprise,
        ).first()
        updated_fields, variant_applied = _enrich_existing_item(
            duka, intp, user, existing, produ_stock, row, variants, image_map,
        )
        if updated_fields:
            parts_swa = ', '.join(updated_fields)
            return {
                'success': True,
                'row': row_num,
                'item_id': produ_stock.id if produ_stock else existing.id,
                'name': name,
                'enriched': True,
                'message_swa': f'Bidhaa "{name}" imesasishwa ({parts_swa})',
                'message_eng': f'Item "{name}" was updated ({parts_swa})',
            }
        return {
            'success': False,
            'row': row_num,
            'message_swa': f'Bidhaa "{name}" tayari ipo',
            'message_eng': f'Item "{name}" already exists',
        }

    vipimo_jum = _cell_str(row.get('vipimo_jum')) or 'pc'
    vipimo_reja = _cell_str(row.get('vipimo_reja')) or vipimo_jum
    uwiano = max(_parse_int(row.get('uwiano'), 1), 1)
    idadi_jum = _parse_int(row.get('idadi_jumla'), 0)
    idadi_rej = _parse_int(row.get('idadi_reja'), 0)
    idadi = idadi_rej + (idadi_jum * uwiano)

    bei_kununua = _parse_decimal(row.get('bei_kununua'), 0)
    bei_kuuza = _parse_decimal(row.get('bei_kuuza'), 0)
    bei_kuuza_jum = _parse_decimal(row.get('bei_kuuza_jum'), 0)
    if bei_kuuza_jum <= 0 and bei_kuuza > 0:
        bei_kuuza_jum = bei_kuuza * uwiano

    is_service = _parse_bool_yn(row.get('huduma'), False)
    color_attr = _normalize_kielelezo(row.get('kielelezo'))
    if variants and not color_attr:
        color_attr = 'Model'

    aina_key = (_cell_str(row.get('aina')).lower(), _cell_str(row.get('kundi')).lower())
    if aina_key in ctx['aina']:
        selected_aina, mahi = ctx['aina'][aina_key]
    else:
        selected_aina, mahi = _resolve_aina(duka, row.get('aina'), row.get('kundi'))
        ctx['aina'][aina_key] = (selected_aina, mahi)

    brand_key = _cell_str(row.get('chapa')).lower()
    if brand_key in ctx['brand']:
        brand = ctx['brand'][brand_key]
    else:
        brand = _resolve_brand(duka, row.get('chapa'))
        ctx['brand'][brand_key] = brand

    supplier_key = _cell_str(row.get('wasambazaji')).lower()
    if supplier_key in ctx['supplier']:
        supplier = ctx['supplier'][supplier_key]
    else:
        supplier = _resolve_supplier(duka.owner.user, row.get('wasambazaji'))
        ctx['supplier'][supplier_key] = supplier

    produ = bidhaa()
    produ.kampuni = brand
    produ.bidhaa_aina = selected_aina
    produ.Mahi = mahi
    produ.idadi_jum = float(uwiano)
    produ.change_date = timezone.now()
    produ.maelezo = _cell_str(row.get('maelezo')) or 'none'
    produ.vipimo = vipimo_reja
    produ.vipimo_jum = vipimo_jum
    produ.bidhaa_jina = name
    produ.saletaxInluded = False
    produ.purchtaxInluded = False
    produ.namba = _cell_str(row.get('namba'))
    produ.material = False
    produ.owner = ctx['owner_user']
    if color_attr or variants:
        produ.colorAttr = color_attr or 'Model'
    produ.save()

    produ_stock = bidhaa_stoku()
    produ_stock.bidhaa = produ
    produ_stock.idadi = idadi
    produ_stock.Interprise = intp.Interprise
    if supplier:
        produ_stock.msambaji = supplier
    produ_stock.Bei_kununua = float(bei_kununua)
    produ_stock.Bei_kuuza = float(bei_kuuza)
    produ_stock.Bei_kuuza_jum = float(bei_kuuza_jum)
    produ_stock.op_name = ctx['op_user']
    produ_stock.expire_date = None
    produ_stock.sirio = _cell_str(row.get('barcode'))
    produ_stock.tanguliziwa = 0
    produ_stock.service = is_service
    produ_stock.timely = 0

    adj_str, adj_num_next = _next_adj_code(duka, ctx['adj_seq'])
    adj = stokAdjustment()
    adj.Interprise = duka
    adj.date = timezone.now()
    adj.code = adj_str
    adj.code_num = adj_num_next
    adj.Na = intp
    adj.registered = True
    if idadi > 0:
        adj.Ongezwa = True
    adj.save()

    produ_stock.ongezwa = adj
    produ_stock.save()

    reg = productChangeRecord()
    reg.prod = produ_stock
    reg.qty = idadi
    reg.adjst = adj
    reg.save()

    if variants:
        variants = _enrich_variants_from_master(row, variants)
        _apply_variants(
            duka, intp, user, produ, produ_stock, reg, uwiano, variants, image_map,
            color_attr=color_attr or produ.colorAttr or '',
        )
    elif _cell_str(row.get('picha')):
        _attach_item_image_no_variant(duka, intp, produ, produ_stock, row.get('picha'), image_map)

    return {
        'success': True,
        'row': row_num,
        'item_id': produ_stock.id,
        'name': name,
        'message_swa': f'Bidhaa "{name}" imeongezwa',
        'message_eng': f'Item "{name}" was added',
    }


def import_item_row(duka, intp, user, row, row_num=0, image_map=None):
    """Backward-compatible single-row import."""
    return import_item_group(
        duka, intp, user,
        {'master': row, 'variants': [_variant_slice(row)] if _has_variant_data(row) else [], 'start_row': row_num},
        image_map=image_map,
        row_num=row_num,
    )


def bulk_import_items(duka, intp, user, rows, image_map=None):
    groups = group_import_rows(rows)
    results = []
    created = 0
    failed = 0
    ctx = _import_ctx(duka, intp, user)
    for group in groups:
        row_num = group.get('start_row', 0)
        try:
            result = import_item_group(duka, intp, user, group, image_map=image_map, row_num=row_num, ctx=ctx)
        except Exception as exc:
            result = {
                'success': False,
                'row': row_num,
                'message_swa': f'Hitilafu mstari {row_num}: {exc}',
                'message_eng': f'Error on row {row_num}: {exc}',
            }
        results.append(result)
        if result.get('success'):
            created += 1
        elif not result.get('skipped'):
            failed += 1
    return {
        'created': created,
        'failed': failed,
        'results': results,
        'groups': len(groups),
    }
